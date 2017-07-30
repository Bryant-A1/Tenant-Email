""" Okay, here's a challenge:
Write a script that reads in data from a text file and an xlsx file, then logs
into my email and sends a bunch of emails.
1, 2, 3 go!
Currently, this is using a TEST file so I don't send a bunch of emails to
actual tenants.

The Excel Workbook is called 'test_ML.xlsx'
"""
# Going to need this package
import openpyxl
# THis package won't work for .ods files, so they have to be in the .xlsx format.
print("Opening Workbook...")
wb = openpyxl.load_workbook('test_ML.xlsx')
# type(wb)
# Self explanatory method
#wb.get_sheet_names()

# Each sheet is represented by a Worksheet object...
# I'll need to change this sheet name later...
sheet = wb.get_sheet_by_name('Sheet1')
# sheet['A1']
# sheet['A1'].value

# You can get a cell using the sheet's 'call()' method and passing integers for its row and column keyword agruments
# The first row or column integer is 1.
# For example:
# sheet.cell(row=1, column=2)

# Let's create some empty lists so we can store tenant data.
'''tenant_Name = []
tenant_Property = []
tenant_lease_end_date = []
tenant_Email = []
'''


class Tenant(object):
    """Tenant class used for clean iterative implementation."""

    def __init__(self, Name, Property, LeaseEndDate, Email):
        """Init Tenant class."""
        self.Name = Name
        self.Property = Property
        self.LeaseEndDate = LeaseEndDate
        self.Email = Email

    def ReturnVals(self):
        """Return values."""
        return self.Name, self.Property, self.LeaseEndDate, self.Email


Tenants = []

def GetInfo():
    for i in range(1, sheet.max_row):
        tName = sheet.cell(row=i, column=1)
        tProp = sheet.cell(row=i, column=2)
        tLease = sheet.cell(row=i, column=3)
        tEmail = sheet.cell(row=i, column=4)
        NewTenant = Tenant(tName, tProp, tLease, tEmail)
        Tenants.append(NewTenant)


GetInfo()
print(Tenants)

"""
# Okay, next step, read in the data from the text file and use formatting to input these values.
# How do I do this...

# Okay now this part is pretty much done... Let's read some data from the text file now.
print("Opening Email Message...")
email_message = open('ConfirmMoveOut.txt')
print("Let's print the message just to be sure")
print(email_message.read())
email_message.seek(0)
input("Are you sure this is the message you want to send? >> ")
str_email_message = str(email_message.read())
email_message.close()

# Let's try printing this email message with the tenant information.
e_messages = []
a = 0
for i in range(1, sheet.max_row):
    e_messages.append(str_email_message.format(tenant_Name[a], tenant_Property[a], tenant_lease_end_date[a]))
    a += 1
# Boom, now we have a list of each message, with the correct tenant name, property, and lease end date. I can display this list of messages by running the line below.
# for message in messages:
#    print(message,'\n')
print("You will now send emails to the following email addresses", tenant_Email)
'''
Next Step: Connect to the email server
'''
import smtplib
import getpass
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
# Now let's establish a connection to the server:
smtpObj.ehlo()
# Now let's enable encryption
smtpObj.starttls()
# Should return something with a value of 220
my_Email = input("What is your email address?")
my_Password = getpass.getpass()

smtpObj.login(my_Email, my_Password)
# Okay, now that we're logged in, it's time to send the email.
a = 0
for email in tenant_Email:
    smtpObj.sendmail(my_Email, tenant_Email[a],
    'Subject: Please Confirm Move Out or Extension\n'+e_messages[a])
    a += 1

smtpObj.quit()
"""
